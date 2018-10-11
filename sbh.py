from measurement import Measurement
from collections import OrderedDict
import xlrd
import openpyxl as xl
import scipy.constants as scc
import numpy as np 

class SBH:
    def __init__(self, Measurements, order=1.5):
        # list of measurements instantiated from the
        # Measurement class
        self.measurements = Measurements    
        self.order = order
        # assume the list is not sorted by temperature and does so
        # doing this in the constructor ensures it is always sorted
        self.__sort_by_temperature()

    def __repr__(self):
        for i in range( len( self.measurements ) ):
            print "[", i, "] = ", self.measurements[i]
        
    def __str__(self):
        return self.__repr__()

    def get_num_measurements(self):
        """ Called check the number of measurements 
        contained in self.measurements
        """
        return len( self.measurements )   

    def add_measurement(self, item):
        """ Called to add a measurement to the 
        list of existing measurements

        Parameters
        ----------
        item : Measurement
        """
        self.measurements.append(item)

    def remove_measurement(self, index=None):
        """ Called to remove a measurement from the
        list of measurements

        Parameters
        ----------
        index : int or None
            index of the measurement to be removed,
            if it is None, then the default is to remove
            the last item
        """ 
        if index is not None:
            return self.measurements.pop(index)
        else:
            return self.measurements.pop()

    def __sort_by_temperature(self):
        """ Called to sort the list of measurements by temperature
        in descending order
        """
        temps = [] 
        # first get each temperature
        for i in range( len( self.measurements ) ):
            temps.append( self.measurements[i].temperature() )
        # reverse sort (descending order)
        self.measurements = [x for _, x in sorted(zip(temps, self.measurements), key=lambda pair: pair[0], reverse=True)]

    def get_temperatures(self):
        """ Called to return an array of all the temperatures
        in self.measurements
        """
        temps = []
        for i in range( len( self.measurements ) ):
            temps.append( self.measurements[i].temperature() )
        return temps
    
    def lnT_current(self, item, index=None):
        """ Called to return the whole lnT array for a given
        measurement

        Parameters
        ----------
        item : int 
            refers to specific measurement 
        index : int or None 
            refers to a specific value within the aforementioned item (measurement)
        """
        try:
            if type(index) == int:
                return self.measurements[item].lnT_current(order=self.order)[index]
            else:
                return self.measurements[item].lnT_current(order=self.order)
        except IndexError:
            print "Index out of Bounds, Measurement index does not exist"
            return None
    
    def current(self, item, index=None):
        """ Called to return whole current array for a given 
        measurement

        Parameters
        ----------
        item : int 
            refers to specific measurement 
        index : int or None 
            refers to a specific value within the aforementioned item (measurement)
        """
        try:
            if type(index) == int:
                return self.measurements[item].current()[index]
            else:
                return self.measurements[item].current()
        except IndexError:
            print "Index out of Bounds, Measurement index does not exist"
            return None 
    
    def __check_gate_step_consistency(self):
        """ Called to check that the gate step size is the same throughout all the 
        measurements in self.measurements, if it is not it will return false and
        the further analysis cannot be completed
        """
        steps = [] 
        for i in range( len( self.measurements ) ):
            # get the steps of each measurement, round for sake of equality checks
            steps.append( round(self.measurements[i].gate_voltage()[1] - self.measurements[i].gate_voltage()[0], 2) )
        # compare all steps to make sure they are equal
        if all( x == steps[0] for x in steps ):
            return True
        else:
            return False
    
    def __generate_output_headers(self):
        """ Called to generate an array of headers for each voltage column
        """
        # check that step sizes are equal
        if self.__check_gate_step_consistency() == True:
            base = "I @ Vgs = " 
            lnT_base = "Ln(I/T^("+str(self.order)+")) @ Vgs = "
            heads = []
            # loop over the first dimension of heads, e.g. len(heads)
            for i in range( len( self.measurements[0].gate_voltage() ) ):
                # round for readability
                val = round( self.measurements[0].gate_voltage()[i], 2 )
                heads.append( base + str(val) + "V" )
                heads.append( lnT_base + str(val) + "V" )
            return heads
        else:
            print "Gate step size is not equal"
            return None

    def __lnT_current_dict(self):
        """ Called to make the current dictionary with the correct headers
        """
        current = [ [] for i in range( len(self.measurements[0].gate_voltage() ) * 2) ]
        # loop over all the measurements
        for i in range( len( self.measurements ) ):
            # loop over each row in the measurement 
            for j in range( len(self.measurements[i].gate_voltage() ) ):
                # fill alternate lists with either current or lnT current
                # j * 2 = 0, 2, 4, 6, ... 
                # j * 2 + 1 = 1, 3, 5, 7, ...
                current[j*2].append( self.current(i, j) ) # i-th measurement, j-th column of that measurement
                current[j*2+1].append( self.lnT_current(i, j) ) 
        return OrderedDict( zip(self.__generate_output_headers(), current) )
    
    def __temperature_dict(self):
        """ Called to create a temperature dict
        with three headers and the corresponding values for each header taken
        from the file
        """
        headers = ["T (K)", "1/kbT (1/eV)", "1000/T (1/K)"]
        # create a list for each header
        t_vals = [ [] for s in range( len( headers ) ) ]
        # loop over all temperatures
        for t in range( len( self.get_temperatures() ) ):
            t_vals[0].append( self.get_temperatures()[t] )
            t_vals[1].append( scc.e / ( scc.k * self.get_temperatures()[t] ) )
            t_vals[2].append( 1000./self.get_temperatures()[t] )
        return OrderedDict( zip( headers, t_vals ) )

    def __voltage_phiB_dict(self):
        """ Called to create a voltage and phib dict
        with the corresponding headers and values
        """
        return ["Vgs (V)", "PhiB (eV)", "PhiB (meV)", "R"]
        
    
    def __merge_dicts(self, *dict_args):
        """ Given any number of dicts, shallow copy and merge into a new
        dicts, precedence goes to key value pairs in latter dicts

        Parameters
        ----------
        *dict_args : dictionary or dictionaries 
            pass any number of dictionaries to be merged
            e.x. merge_dicts( a, b, c, d, ..., z )
        """
        # use ordered dictionaries to preserve key, value order
        result = OrderedDict() 
        for dictionary in dict_args:
            result.update( dictionary )
        return result

    def write_analysis(self, save_name="SBH_analysis.xls"):
        """ Called to write the analysis 
        to two excel sheets
        1. Current vs. Temperature
            contains:
            T, 1/kT, 1000/T, I @ Vgs= xxV, ln(I/T^(3/2)) @ Vgs= xxV,... 
        2. Voltage vs. PhiB
            contains:
            Vgs, PhiB(eV), PhiB(meV), R
        
        Parameters
        ----------
        save_name : string
            name to save excel sheet to
        """
        # create new workbook
        wb = xl.Workbook()
        # create new worksheet
        data_sheet = wb.create_sheet('Current vs. Temperature')
        voltage_sheet= wb.create_sheet("Voltage vs. PhiB")
        # merge and generate data for the sheet
        data_dict = self.__merge_dicts( self.__temperature_dict(), self.__lnT_current_dict() )
        headers = list( data_dict.keys() )
        voltage_headers = self.__voltage_phiB_dict()
        dims = {}
        # fill the data sheet
        for i in range( 1, len( headers ) + 1):
            data_sheet.cell(column=i, row=1, value=headers[i-1])
        count = 1
        for val in data_dict.itervalues():
            for i in range(2):
                data_sheet.cell(column=count, row=i+2, value=val[i])
            count += 1
        # fill the voltage sheet
        for i in range( 1, len( voltage_headers ) + 1 ):
            voltage_sheet.cell(column=i, row=1, value=voltage_headers[i-1])
        for i in range( 2, len( self.measurements[0].gate_voltage() ) ):
            voltage_sheet.cell(column=1, row=i, value=self.measurements[0].gate_voltage()[i-2])
        # set the default column width to be the size of the string in the first row
        # for the data sheet
        for row in data_sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            data_sheet.column_dimensions[col].width = value
        # set the default column width to be the size of the string in the first row
        # for the voltage sheet
        for row in voltage_sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            voltage_sheet.column_dimensions[col].width = value
        wb.save(save_name)
        wb.close()