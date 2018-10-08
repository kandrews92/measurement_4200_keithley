#!/usr/bin/env python

def list_dir( ext=".xls", path="." ):
    """ Called to get the files in the current path for analysis
    and return a list of files and their associated temperatures
    sorted by descending order

    Parameters
    ----------
    ext : string
        file extension to be searched for. Should generally be either
        .xls or .xlsx
    path : string
        the file path to search for files
    """
    files = [] 
    T = [] 
    for file in os.listdir( path ):
        # check for file extension and if the file also has the 
        # string 'vgs-id' in it, which is the proper prefix for 
        # measurement files
        if file.endswith( ext ) and "vgs-id" in file:
            files.append( file )
            T.append( get_temperature( file ) )
    # sort the temperatures and the files into a tupled list
    # so that they can be associated to each other and sorted 
    # in descending order
    combined_list = zip( *sorted( zip( T, files ), reverse=True ) )
    T = list( combined_list[0] )       
    files = list( combined_list[1] )
    return files, T

def get_temperature( file ):
    """ Called to get the temperature from a file name

    Parameters
    ----------
    file : string
        for example "vgs-id#2_vds_1V_vgs_-50_80V_vac07_220K.xls
    """
    # split the file string at each of "_" and make each split 
    # element part of a list
    temp = file.split('_') 
    # the last element of the split list will be for example, 220K.xls
    # split this element at the "." and save the 220K as the variable
    temp = temp[-1].split('.')[0] 
    # return the temperature string with the "K" removed and convert
    # the remaining value to a float
    return float(temp[:-1])

def make_current_voltage_dict( T, files ):      
    headers = [] 
    gateV = [] 
    nrows, ncols = get_row_cols( files[0] )
    current = [ [] for i in range( nrows * 2 - 2) ]
    count = 1 
    for f in range( len( files) ):
        wb = xlrd.open_workbook( files[f] )
        sheet = wb.sheet_by_index( 0 ) # data sheet only
        for col in range( ncols ):
            for row in range( 1, sheet.nrows ):
                if (sheet.cell_value( 0, col ) == "GateV") and (count == 1):
                    gateV.append( sheet.cell_value( row, col ) )
                if sheet.cell_value( 0, col ) == "DrainI":
                    val = sheet.cell_value( row, col )
                    try:
                        ln_T = np.log( val/ ( sc.k * T[f]**(3./2.) ) )
                    except RuntimeWarning:
                        ln_T = 0.0
                    current[ row * 2 - 2 ].append( val ) # current (A)
                    current[ row * 2 - 1 ].append( ln_T ) 
        count += 1   
    I_base_str = "I @ V_gs = " 
    lnT_base_str = "ln(I/T^(3/2)) @ V_gs = "
    for val in range( 0, len( gateV ) ):
        I_header = I_base_str + str( round(gateV[val], 2)) + " V"
        lnT_header = lnT_base_str + str( round(gateV[val], 2)) + " V"
        headers.append( I_header )
        headers.append( lnT_header )
    return OrderedDict( zip( headers, current ) ), gateV

def merge_dicts( *dict_args ):
    """ Given any number of dicts, shallow copy and merge into a new
    dicts, precedence goes to key value pairs in latter dicts

    Parameters
    ----------
    *dict_args : dictionary or dictionaries 
        pass any number of dictionaries to be merged
        e.x. merge_dicts( a, b, c, d, ..., z )
    """
    # use ordered dictionaries to preserve key, value order
    result = OrderedDict()
    for dictionary in dict_args:
        result.update( dictionary )
    return result

def make_temperature_dict( T ):
    headers = ["T (K)", "1/kbT (1/eV)", "1000/T (1/K)"]
    T_vals = [ [] for i in range( len(headers) ) ]
    for t in range( len( T ) ):
        T_vals[0].append( T[t] )
        T_vals[1].append( sc.e/(sc.k * T[t] ) )
        T_vals[2].append( 1000./T[t] )
    return OrderedDict( zip( headers, T_vals) )

def get_row_cols( file ):
    """ Called to get the number of rows and columns of an excel file

    Parameters
    ----------
    file : string
        excel file to be opened
    """
    wb = xlrd.open_workbook( file )
    sheet = wb.sheet_by_index( 0 )
    return sheet.nrows, sheet.ncols

def write_data_excel( all_dict, gateV, output_path="output/", fname="" ):
    wb = openpyxl.Workbook()
    data_sheet = wb.create_sheet("Current vs Temperature")
    voltage_sheet = wb.create_sheet("Voltage vs. PhiB")
    v_sheet_heads = ["Vgs (V)", "PhiB (eV)", "PhiB (meV)", "R"]
    headers = list( all_dict.keys() )
    dims = {}
    for i in range(1,len(headers)+1):
        data_sheet.cell(column=i, row=1, value=headers[i-1])
    count = 1
    for val in all_dict.itervalues():
        for i in range(2):
            data_sheet.cell(column=count, row=i+2, value=val[i])
        count += 1
    for i in range( 1, len(v_sheet_heads) + 1):
        voltage_sheet.cell(column=i, row=1, value=v_sheet_heads[i-1])
    for i in range( 2, len(gateV)+1 ):
        voltage_sheet.cell(column=1, row=i, value=gateV[i-2])
    # set the default column width to be the size of the string in 
    # the first row
    for row in data_sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)))) 
    for col, value in dims.items():
        data_sheet.column_dimensions[col].width = value
    for row in voltage_sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)))) 
    for col, value in dims.items():
        voltage_sheet.column_dimensions[col].width = value
    if not os.path.exists( output_path ):
        os.makedirs( output_path )
    fname = time.strftime("%Y%m%d-%H%M%S")
    wb.save( output_path + "barrier_height_" + fname + ".xls" )
    wb.close()

if __name__ == "__main__":
    import os
    import xlrd # reading excel files
    import openpyxl # writing to large (> 256 cols) excel sheets
    # sc.k = boltzmann const.; sc.e = electron charge
    import scipy.constants as sc
    import numpy as np 
    import time
    from collections import OrderedDict
    
    files, T = list_dir( )
    T_dict = make_temperature_dict( T )
    Data_dict, gateV = make_current_voltage_dict( T, files )
    all_dict = merge_dicts( T_dict, Data_dict )
    write_data_excel( all_dict, gateV )
